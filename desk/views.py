from __future__ import annotations

import base64
import csv
import io
import json
from collections.abc import Callable
from functools import wraps
from typing import Any

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist, PermissionDenied
from django.core.paginator import Paginator
from django.http import Http404, HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_http_methods, require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tvs_dms.exporter import _safe_cell, tabular
from tvs_dms.forms import MODULES, MODULES_BY_ROLE, ROLE_LABELS, modules_for_role
from tvs_dms.security import SecurityError, encrypt_bytes

from .analytics import (
    TIER_LABELS,
    build_report,
    can_see,
    catalogue_for,
    detail_rows,
    parse_window,
    report_rows,
    select_observations,
    tier_of,
)
from .analytics.access import is_identity_field
from .analytics.pdf import render_report_pdf
from .analytics.visibility import redact_row
from .crypto import data_key, key_fingerprint
from .forms import (
    ActivityEntryForm,
    LoginForm,
    PasswordResetForm,
    SetupForm,
    StrongPasswordChangeForm,
    UserCreateForm,
)
from .models import ActivityRecord
from .store import AlreadyInitialized, DuplicateUsername, get_store


def initialized() -> bool:
    return get_store().initialized()


def role_of(user) -> str:
    try:
        return user.desk_profile.role
    except (AttributeError, ObjectDoesNotExist) as exc:
        raise PermissionDenied("This account has no Activity Desk role.") from exc


def admin_required(view: Callable) -> Callable:
    @wraps(view)
    @login_required
    def wrapped(request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        if role_of(request.user) != "administrator":
            raise PermissionDenied
        return view(request, *args, **kwargs)

    return wrapped


def audit(user, action: str, target: str = "") -> None:
    get_store().audit(user, action, target)


def module_enabled(module_key: str) -> bool:
    return get_store().module_states().get(module_key, True)


def allowed_modules(user, include_disabled: bool = False):
    modules = modules_for_role(role_of(user))
    if include_disabled or role_of(user) == "administrator":
        return modules
    disabled = {key for key, enabled in get_store().module_states().items() if not enabled}
    return [module for module in modules if module.key not in disabled]


def records_for(user) -> list:
    records = get_store().list_records()
    return records if role_of(user) == "administrator" else [record for record in records if str(record.owner_id) == str(user.id)]


def record_to_dict(record: ActivityRecord, role: str | None = None) -> dict[str, Any]:
    profile = getattr(record.owner, "desk_profile", None)
    values = record.get_data()
    # The visibility console is deny-only but must bind on every path out of the
    # store, not just reports; an unredacted export would reinstate a field the
    # school deliberately hid.
    if role is not None:
        values = redact_row(values, record.module_key, role)
    return {
        "id": str(record.id),
        "module_key": record.module_key,
        "module_name": record.module_name,
        "role": record.role,
        "status": record.status,
        "owner_name": profile.display_name if profile else record.owner.username,
        "created_at": record.created_at.isoformat(),
        "updated_at": record.updated_at.isoformat(),
        "data": values,
    }


def _sign_in(request: HttpRequest, user) -> None:
    request.session.flush()
    request.session["tvs_user_id"] = str(user.id)
    request.session["tvs_session_version"] = getattr(user.desk_profile, "session_version", 1)


def _sign_out(request: HttpRequest) -> None:
    request.session.flush()


@require_GET
def home(request: HttpRequest) -> HttpResponse:
    if not initialized():
        return redirect("setup")
    return redirect("dashboard" if request.user.is_authenticated else "login")


@require_http_methods(["GET", "POST"])
def setup(request: HttpRequest) -> HttpResponse:
    if initialized():
        return redirect("dashboard" if request.user.is_authenticated else "login")
    form = SetupForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        try:
            user = get_store().create_initial_admin(
                school_name=form.cleaned_data["school_name"].strip(),
                username=form.cleaned_data["username"],
                password=form.cleaned_data["password"],
                display_name=form.cleaned_data["display_name"].strip(),
            )
        except (AlreadyInitialized, DuplicateUsername):
            messages.error(request, "Setup has already been completed. Sign in instead.")
            return redirect("login")
        _sign_in(request, user)
        messages.success(request, "Secure workspace created. Add staff accounts from Users.")
        return redirect("dashboard")
    return render(request, "desk/setup.html", {"form": form, "setup_mode": True})


@require_http_methods(["GET", "POST"])
def login_view(request: HttpRequest) -> HttpResponse:
    if not initialized():
        return redirect("setup")
    if request.user.is_authenticated:
        return redirect("dashboard")
    form = LoginForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        username = form.cleaned_data["username"].strip()
        user = get_store().get_user_by_username(username)
        generic_error = "Invalid username or password."
        if not user or not user.is_active:
            form.add_error(None, generic_error)
        else:
            profile = user.desk_profile
            now = timezone.now()
            if profile.locked_until and profile.locked_until > now:
                seconds = max(1, int((profile.locked_until - now).total_seconds()))
                form.add_error(None, f"Account temporarily locked. Try again in {seconds} seconds.")
            elif not user.check_password(form.cleaned_data["password"]):
                get_store().register_login_failure(user)
                form.add_error(None, generic_error)
            else:
                get_store().clear_login_failures(user)
                _sign_in(request, user)
                audit(user, "login", user.username)
                next_url = request.POST.get("next", "")
                if next_url.startswith("/") and not next_url.startswith("//"):
                    return redirect(next_url)
                return redirect("dashboard")
    return render(request, "desk/login.html", {"form": form, "next": request.GET.get("next", "")})


@require_POST
def logout_view(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        audit(request.user, "logout", request.user.username)
    _sign_out(request)
    return redirect("login")


@login_required
@require_http_methods(["GET", "POST"])
def change_password(request: HttpRequest) -> HttpResponse:
    form = StrongPasswordChangeForm(request.user, request.POST or None)
    if request.method == "POST" and form.is_valid():
        get_store().update_password(request.user, form.cleaned_data["new_password1"], must_change=False)
        request.session["tvs_session_version"] = getattr(request.user.desk_profile, "session_version", 1)
        audit(request.user, "password_changed", request.user.username)
        messages.success(request, "Your password has been changed.")
        return redirect("dashboard")
    return render(request, "desk/change_password.html", {"form": form})


@login_required
def dashboard(request: HttpRequest) -> HttpResponse:
    query = request.GET.get("q", "").strip()
    modules = allowed_modules(request.user)
    if query:
        modules = [module for module in modules if query.casefold() in module.name.casefold()]
    records = records_for(request.user)
    counts = {
        "total": len(records),
        "submitted": sum(record.status == ActivityRecord.SUBMITTED for record in records),
        "draft": sum(record.status == ActivityRecord.DRAFT for record in records),
    }
    grouped: list[tuple[str, list[Any]]] = []
    for role, label in ROLE_LABELS.items():
        role_modules = [module for module in modules if module.role == role]
        if role_modules:
            grouped.append((label, role_modules))
    return render(request, "desk/dashboard.html", {"counts": counts, "module_groups": grouped, "query": query})


def _load_record_for_user(user, record_id: str):
    record = get_store().get_record(record_id)
    if record is None:
        raise Http404("Record not found")
    if role_of(user) != "administrator" and str(record.owner_id) != str(user.id):
        raise PermissionDenied
    return record


@login_required
@require_http_methods(["GET", "POST"])
def record_form(request: HttpRequest, module_key: str, record_id: str | None = None) -> HttpResponse:
    module = MODULES.get(module_key)
    if module is None:
        raise Http404("Unknown form")
    user_role = role_of(request.user)
    if user_role != "administrator" and user_role != module.role:
        raise PermissionDenied
    if user_role != "administrator" and not module_enabled(module.key):
        messages.error(request, "This form has been disabled by an administrator.")
        return redirect("dashboard")

    record = _load_record_for_user(request.user, record_id) if record_id else None
    if record and record.module_key != module.key:
        raise Http404("Record does not belong to this form")
    try:
        initial = record.get_data() if record else {}
    except SecurityError:
        messages.error(request, "This record could not be decrypted. Ask the administrator to check the data key.")
        return redirect("records")

    action = request.POST.get("action", "draft")
    submitted = action == "submit"
    form = ActivityEntryForm(
        request.POST or None,
        module=module,
        submitted=submitted,
        initial=initial,
    )
    if request.method == "POST" and form.is_valid():
        payload = form.payload()
        record = get_store().save_record(
            record=record,
            module_key=module.key,
            module_name=module.name,
            role=module.role,
            owner=request.user,
            status=ActivityRecord.SUBMITTED if submitted else ActivityRecord.DRAFT,
            event_date=form.cleaned_data.get("event_date"),
            payload=payload,
        )
        audit(request.user, "record_updated" if record_id else "record_created", f"{module.key}:{record.id}")
        messages.success(request, f"{module.name} saved as {'submitted' if submitted else 'a draft'}.")
        return redirect("records")
    return render(
        request,
        "desk/record_form.html",
        {"form": form, "module": module, "module_role_label": ROLE_LABELS[module.role], "record": record},
    )


@login_required
def records(request: HttpRequest) -> HttpResponse:
    record_rows = records_for(request.user)
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    module_key = request.GET.get("module", "").strip()
    if query:
        folded = query.casefold()
        record_rows = [
            record
            for record in record_rows
            if folded in record.module_name.casefold()
            or folded in record.owner.username.casefold()
            or folded in record.owner.desk_profile.display_name.casefold()
        ]
    if status in {ActivityRecord.DRAFT, ActivityRecord.SUBMITTED}:
        record_rows = [record for record in record_rows if record.status == status]
    available = allowed_modules(request.user, include_disabled=True)
    allowed_keys = {module.key for module in available}
    if module_key in allowed_keys:
        record_rows = [record for record in record_rows if record.module_key == module_key]
    page = Paginator(record_rows, 40).get_page(request.GET.get("page"))
    return render(
        request,
        "desk/records.html",
        {"page": page, "modules": available, "query": query, "status": status, "selected_module": module_key},
    )


@login_required
def reports(request: HttpRequest) -> HttpResponse:
    """Report index: every module this user can analyse, with its access tier."""
    role = role_of(request.user)
    modules = allowed_modules(request.user, include_disabled=True)
    return render(
        request,
        "desk/reports.html",
        {
            "modules": modules,
            "tier": tier_of(role),
            "tier_label": TIER_LABELS.get(tier_of(role), ""),
            "windows": REPORT_WINDOWS,
        },
    )


REPORT_WINDOWS = (
    ("30", "Last 30 days"),
    ("90", "Last 90 days"),
    ("180", "Last 6 months"),
    ("365", "Last 12 months"),
    ("all", "All records"),
)


def _report_context(request: HttpRequest, module_key: str):
    """Shared by the HTML report view and every export of it."""
    role = role_of(request.user)
    allowed = {module.key: module for module in allowed_modules(request.user, include_disabled=True)}
    module = allowed.get(module_key)
    if module is None:
        raise Http404

    window = request.GET.get("window", "90")
    standard = request.GET.get("standard", "")
    since, period_label = parse_window(window)
    observations = select_observations(
        get_store().list_records(),
        module,
        viewer_id=str(request.user.id),
        tier=tier_of(role),
        since=since,
        standard=standard,
    )
    report = build_report(module, observations, role=role, period_label=period_label)
    return module, report, window, standard


@login_required
@require_GET
def module_report(request: HttpRequest, module_key: str) -> HttpResponse:
    module, report, window, standard = _report_context(request, module_key)
    standards = sorted(
        {
            str(field_choice)
            for field in module.fields
            if field.key == "standard"
            for field_choice in field.choices
        }
    )
    return render(
        request,
        "desk/module_report.html",
        {
            "module": module,
            "report": report,
            "window": window,
            "standard": standard,
            "standards": standards,
            "windows": REPORT_WINDOWS,
            "tier_label": TIER_LABELS.get(report.tier, ""),
            "is_admin": role_of(request.user) == "administrator",
        },
    )


@login_required
@require_GET
def export_report(request: HttpRequest, module_key: str, file_type: str) -> HttpResponse:
    module, report, _window, _standard = _report_context(request, module_key)
    timestamp = timezone.localtime().strftime("%Y%m%d-%H%M")
    slug = module.key.replace("_", "-")
    headers, rows = report_rows(report)

    if file_type == "csv":
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows([[_safe_cell(cell) for cell in row] for row in rows])
        response = HttpResponse("﻿" + output.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="TVS-{slug}-report-{timestamp}.csv"'
    elif file_type == "xlsx":
        response = _report_workbook(report, headers, rows, slug, timestamp)
    elif file_type == "pdf":
        if role_of(request.user) != "administrator":
            raise PermissionDenied("PDF export is restricted to administrators.")
        payload = render_report_pdf(
            report,
            get_store().school_name(),
            timezone.localtime().strftime("%d %b %Y %H:%M"),
            ROLE_LABELS.get(report.role, report.role),
        )
        response = HttpResponse(payload, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="TVS-{slug}-report-{timestamp}.pdf"'
    else:
        raise Http404

    audit(request.user, f"export_report_{file_type}", f"{module.key}: {len(rows)} analyses")
    return response


def _report_workbook(report, headers, rows, slug: str, timestamp: str) -> HttpResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Insights"
    _write_sheet(sheet, headers, rows)
    for title, block_headers, block_rows in detail_rows(report):
        name = _sheet_name(title, {ws.title for ws in workbook.worksheets})
        _write_sheet(workbook.create_sheet(name), block_headers, block_rows)
    output_bytes = io.BytesIO()
    workbook.save(output_bytes)
    response = HttpResponse(
        output_bytes.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="TVS-{slug}-report-{timestamp}.xlsx"'
    return response


def _sheet_name(title: str, taken: set[str]) -> str:
    # Excel sheet names: 31 chars, and none of : \ / ? * [ ]
    cleaned = "".join(character for character in title if character not in ":\\/?*[]")[:28] or "Detail"
    candidate, suffix = cleaned, 2
    while candidate in taken:
        candidate = f"{cleaned[:26]} {suffix}"
        suffix += 1
    return candidate


def _write_sheet(sheet, headers, rows) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(max(1, len(headers)))}1"
    fill = PatternFill("solid", fgColor="183B66")
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(1, column, str(header))
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    for row_number, row in enumerate(rows, 2):
        for column, value in enumerate(row, 1):
            sheet.cell(row_number, column, _safe_cell(value))
    for column, header in enumerate(headers, 1):
        lengths = [len(str(row[column - 1])) for row in rows[:200] if column - 1 < len(row)]
        sheet.column_dimensions[get_column_letter(column)].width = min(
            60, max([12, len(str(header)) + 2] + lengths)
        )


@admin_required
@require_http_methods(["GET", "POST"])
def field_visibility(request: HttpRequest) -> HttpResponse:
    """Super-admin console: pick a form, pick a role, toggle fields on and off.

    Deny-only by construction — a toggle can hide a field the tier matrix
    already allows, never reveal one it withholds.
    """
    store = get_store()
    modules = list(MODULES.values())
    module_key = request.GET.get("module") or request.POST.get("module_key") or modules[0].key
    module = MODULES.get(module_key)
    if module is None:
        raise Http404
    role = request.GET.get("role") or request.POST.get("role") or module.role
    if role not in ROLE_LABELS:
        raise Http404

    if request.method == "POST":
        field_key = request.POST.get("field_key", "")
        if field_key not in {field.key for field in module.fields}:
            raise Http404
        hidden = request.POST.get("hidden") == "true"
        store.set_field_hidden(module.key, role, field_key, hidden)
        audit(
            request.user,
            "field_visibility",
            f"{module.key}/{role}/{field_key} -> {'hidden' if hidden else 'visible'}",
        )
        messages.success(
            request,
            f"{'Hidden' if hidden else 'Restored'} “{field_key}” for {ROLE_LABELS[role]}.",
        )
        return redirect(f"{reverse('field_visibility')}?module={module.key}&role={role}")

    hidden_keys = store.hidden_fields().get((module.key, role), set())
    tier = tier_of(role)
    rows = [
        {
            "field": field,
            "hidden": field.key in hidden_keys,
            "identity": is_identity_field(field.key),
            "blocked_by_tier": tier == "decision" and is_identity_field(field.key),
        }
        for field in module.fields
    ]
    catalogue = [entry for entry in catalogue_for(module) if can_see(tier, entry.tier)]
    hidden_analyses = sum(
        1 for entry in catalogue if set(entry.fields_used) & hidden_keys
    )

    return render(
        request,
        "desk/field_visibility.html",
        {
            "modules": modules,
            "module": module,
            "roles": [(key, label) for key, label in ROLE_LABELS.items()],
            "role": role,
            "role_label": ROLE_LABELS[role],
            "tier": tier,
            "tier_label": TIER_LABELS.get(tier, ""),
            "rows": rows,
            "hidden_count": len(hidden_keys),
            "analysis_count": len(catalogue),
            "hidden_analyses": hidden_analyses,
        },
    )


def _filtered_export_records(request: HttpRequest) -> list[dict[str, Any]]:
    record_rows = records_for(request.user)
    status = request.GET.get("status", "")
    module_key = request.GET.get("module", "")
    allowed_keys = {module.key for module in allowed_modules(request.user, include_disabled=True)}
    if status in {ActivityRecord.DRAFT, ActivityRecord.SUBMITTED}:
        record_rows = [record for record in record_rows if record.status == status]
    if module_key in allowed_keys:
        record_rows = [record for record in record_rows if record.module_key == module_key]
    role = role_of(request.user)
    return [record_to_dict(record, role) for record in record_rows]


@login_required
@require_GET
def export_records(request: HttpRequest, file_type: str) -> HttpResponse:
    rows = _filtered_export_records(request)
    headers, values = tabular(rows)
    timestamp = timezone.localtime().strftime("%Y%m%d-%H%M")
    if file_type == "csv":
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(values)
        response = HttpResponse("\ufeff" + output.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="TVS-activity-records-{timestamp}.csv"'
    elif file_type == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Activity Records"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(max(1, len(headers)))}1"
        fill = PatternFill("solid", fgColor="183B66")
        for column, header in enumerate(headers, 1):
            cell = sheet.cell(1, column, header)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
        for row_number, row in enumerate(values, 2):
            for column, value in enumerate(row, 1):
                sheet.cell(row_number, column, value)
        for column, header in enumerate(headers, 1):
            lengths = [len(str(row[column - 1])) for row in values[:200]]
            sheet.column_dimensions[get_column_letter(column)].width = min(45, max([12, len(header) + 2] + lengths))
        output_bytes = io.BytesIO()
        workbook.save(output_bytes)
        response = HttpResponse(
            output_bytes.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="TVS-activity-records-{timestamp}.xlsx"'
    else:
        raise Http404
    audit(request.user, f"export_{file_type}", f"{len(rows)} records")
    return response


@admin_required
@require_http_methods(["GET", "POST"])
def users(request: HttpRequest) -> HttpResponse:
    form = UserCreateForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        try:
            user = get_store().create_user(
                username=form.cleaned_data["username"],
                password=form.cleaned_data["password"],
                display_name=form.cleaned_data["display_name"].strip(),
                role=form.cleaned_data["role"],
            )
        except DuplicateUsername:
            form.add_error("username", "That username already exists.")
        else:
            audit(request.user, "user_created", user.username)
            messages.success(request, f"Account created for {user.desk_profile.display_name}.")
            return redirect("users")
    user_rows = get_store().list_users()
    return render(request, "desk/users.html", {"form": form, "users": user_rows})


@admin_required
@require_POST
def toggle_user(request: HttpRequest, user_id: str) -> HttpResponse:
    target = get_store().get_user(user_id)
    if target is None:
        raise Http404("User not found")
    if target == request.user:
        messages.error(request, "You cannot deactivate your own account.")
    else:
        get_store().set_user_active(target, not target.is_active)
        action = "user_activated" if target.is_active else "user_deactivated"
        audit(request.user, action, target.username)
        messages.success(request, f"{target.desk_profile.display_name} is now {'active' if target.is_active else 'inactive'}.")
    return redirect("users")


@admin_required
@require_http_methods(["GET", "POST"])
def reset_password(request: HttpRequest, user_id: str) -> HttpResponse:
    target = get_store().get_user(user_id)
    if target is None:
        raise Http404("User not found")
    form = PasswordResetForm(request.POST or None, user=target)
    if request.method == "POST" and form.is_valid():
        get_store().update_password(target, form.cleaned_data["password"], must_change=target != request.user)
        audit(request.user, "password_reset", target.username)
        messages.success(request, f"Password reset for {target.desk_profile.display_name}.")
        return redirect("users")
    return render(request, "desk/reset_password.html", {"form": form, "target": target})


@admin_required
@require_http_methods(["GET", "POST"])
def form_controls(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        module_key = request.POST.get("module_key", "")
        module = MODULES.get(module_key)
        if not module:
            raise Http404
        enabled = request.POST.get("enabled") == "true"
        get_store().set_module_enabled(module_key, enabled)
        audit(request.user, "form_enabled" if enabled else "form_disabled", module_key)
        messages.success(request, f"{module.name} {'enabled' if enabled else 'disabled'}.")
        return redirect(f"{reverse('form_controls')}#{module_key}")
    states = get_store().module_states()
    groups = [
        (ROLE_LABELS[role], [(module, states.get(module.key, True)) for module in modules])
        for role, modules in MODULES_BY_ROLE.items()
    ]
    return render(request, "desk/form_controls.html", {"module_groups": groups})


@admin_required
def audit_log(request: HttpRequest) -> HttpResponse:
    page = Paginator(get_store().list_audit(), 100).get_page(request.GET.get("page"))
    return render(request, "desk/audit.html", {"page": page})


@admin_required
@require_GET
def backup(request: HttpRequest) -> HttpResponse:
    return render(request, "desk/backup.html", {"fingerprint": key_fingerprint()})


@admin_required
@require_POST
def backup_download(request: HttpRequest) -> HttpResponse:
    users_data = []
    for user in get_store().list_users():
        users_data.append(
            {
                "username": user.username,
                "password_hash": user.password,
                "active": user.is_active,
                "display_name": user.desk_profile.display_name,
                "role": user.desk_profile.role,
                "must_change_password": user.desk_profile.must_change_password,
            }
        )
    records_data = []
    for record in get_store().list_records():
        records_data.append(
            {
                "id": str(record.id),
                "module_key": record.module_key,
                "module_name": record.module_name,
                "role": record.role,
                "owner": record.owner.username,
                "status": record.status,
                "event_date": record.event_date.isoformat() if record.event_date else "",
                "nonce": base64.b64encode(bytes(record.payload_nonce)).decode("ascii"),
                "ciphertext": base64.b64encode(bytes(record.payload_ciphertext)).decode("ascii"),
                "tag": base64.b64encode(bytes(record.payload_tag)).decode("ascii"),
                "created_at": record.created_at.isoformat(),
                "updated_at": record.updated_at.isoformat(),
            }
        )
    payload = {
        "format": "tvs-cloud-backup",
        "version": 1,
        "created_at": timezone.now().isoformat(),
        "key_fingerprint": key_fingerprint(),
        "school_name": get_store().school_name(),
        "users": users_data,
        "module_controls": [
            {"module_key": module_key, "enabled": enabled}
            for module_key, enabled in get_store().module_states().items()
        ],
        "records": records_data,
    }
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    nonce, ciphertext, tag = encrypt_bytes(raw, data_key(), b"tvs-cloud-backup-v1")
    envelope = {
        "format": "tvs-cloud-envelope",
        "version": 1,
        "nonce": base64.b64encode(nonce).decode("ascii"),
        "ciphertext": base64.b64encode(ciphertext).decode("ascii"),
        "tag": base64.b64encode(tag).decode("ascii"),
    }
    content = json.dumps(envelope, separators=(",", ":")).encode("utf-8")
    audit(request.user, "backup_created", f"{len(records_data)} records")
    filename = f"TVS-backup-{timezone.localtime():%Y%m%d-%H%M}.tvsbackup"
    response = HttpResponse(content, content_type="application/octet-stream")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@require_GET
def health(request: HttpRequest) -> JsonResponse:
    get_store().health()
    return JsonResponse({"ok": True, "service": "tvs-activity-desk"})
