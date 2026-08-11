from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

from .forms import MODULES


META_COLUMNS = (
    ("record_id", "Record ID"),
    ("module_name", "Form"),
    ("role", "Role"),
    ("status", "Status"),
    ("owner_name", "Entered by"),
    ("created_at", "Created at"),
    ("updated_at", "Updated at"),
)


def _safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).replace("\x00", "")
    if text.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def tabular(records: Sequence[Dict[str, Any]]) -> Tuple[List[str], List[List[Any]]]:
    field_keys: List[str] = []
    labels: Dict[str, str] = {}
    for record in records:
        module = MODULES.get(record.get("module_key", ""))
        if module:
            labels.update({field.key: field.label for field in module.fields})
        for key in record.get("data", {}):
            if key not in field_keys:
                field_keys.append(key)
    headers = [label for _, label in META_COLUMNS] + [labels.get(key, key.replace("_", " ").title()) for key in field_keys]
    rows: List[List[Any]] = []
    for record in records:
        meta = {
            "record_id": record.get("id", ""),
            "module_name": record.get("module_name", ""),
            "role": record.get("role", "").replace("_", " ").title(),
            "status": record.get("status", "").title(),
            "owner_name": record.get("owner_name", ""),
            "created_at": record.get("created_at", "").replace("T", " "),
            "updated_at": record.get("updated_at", "").replace("T", " "),
        }
        data = record.get("data", {})
        rows.append([_safe_cell(meta[key]) for key, _ in META_COLUMNS] + [_safe_cell(data.get(key, "")) for key in field_keys])
    return headers, rows


def export_csv(records: Sequence[Dict[str, Any]], destination: Path) -> Path:
    headers, rows = tabular(records)
    destination = Path(destination)
    with destination.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)
    return destination


def export_xlsx(records: Sequence[Dict[str, Any]], destination: Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("XLSX support is not installed. Run: pip install openpyxl") from exc

    headers, rows = tabular(records)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Activity Records"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:%s1" % get_column_letter(max(1, len(headers)))
    header_fill = PatternFill("solid", fgColor="183B66")
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column, header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    for row_number, values in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, value)
    for column, header in enumerate(headers, start=1):
        sample = [len(str(row[column - 1])) for row in rows[:200]]
        sheet.column_dimensions[get_column_letter(column)].width = min(45, max([len(header) + 2, 12] + sample))
    sheet.row_dimensions[1].height = 24
    workbook.save(str(destination))
    return destination

